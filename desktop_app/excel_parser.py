import openpyxl
import logging
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

class ExcelParser:
    """
    Utility class to read Excel spreadsheets and map columns cross-platform.
    """
    @staticmethod
    def get_headers(file_path: str) -> list[str]:
        """
        Loads the Excel workbook and retrieves headers from the first active row.
        Resolves path using pathlib for cross-platform safety.
        """
        try:
            resolved_path = Path(file_path).resolve()
            wb = openpyxl.load_workbook(resolved_path, read_only=True, data_only=True)
            sheet = wb.active
            if not sheet:
                return []
            
            # Read first row for headers
            for row in sheet.iter_rows(max_row=1, values_only=True):
                if row:
                    # Clean up header values (convert to string, strip whitespace, remove None)
                    headers = [str(cell).strip() for cell in row if cell is not None]
                    return headers
            return []
        except Exception as e:
            logging.error(f"Error reading headers from {file_path}: {e}")
            raise e

    @staticmethod
    def parse_records(file_path: str) -> list[dict]:
        """
        Parses rows from Excel file automatically assuming:
        - Column 1 (index 0): Barcode
        - Column 2 (index 1): Product Name
        - Column 3 (index 2): Price
        Starts iterating from row 1 (no headers expected).
        Returns a list of dicts: [{"barcode": "...", "product_name": "...", "price_iqd": 0}]
        """
        try:
            resolved_path = Path(file_path).resolve()
            wb = openpyxl.load_workbook(resolved_path, read_only=True, data_only=True)
            sheet = wb.active
            if not sheet:
                return []

            records = []
            row_count = 0
            
            # Start iterating from row 1 (no header row)
            for row in sheet.iter_rows(min_row=1, values_only=True):
                # Ensure row is not entirely empty
                if not row or not any(cell is not None for cell in row):
                    continue
                
                # Extract values by index safely
                barcode_val = row[0] if len(row) > 0 else None
                name_val = row[1] if len(row) > 1 else None
                price_val = row[2] if len(row) > 2 else None

                # Clean barcode
                if barcode_val is None:
                    continue
                
                # Check if it is a numeric float or int, convert to clean integer string to avoid scientific notation
                if isinstance(barcode_val, (int, float)):
                    try:
                        barcode_str = str(int(round(float(barcode_val))))
                    except (ValueError, TypeError):
                        barcode_str = str(barcode_val).strip().split('.')[0]
                else:
                    barcode_str = str(barcode_val).strip().split('.')[0]
                    
                if not barcode_str or barcode_str == "None":
                    continue

                # Clean product name
                name_str = str(name_val).strip() if name_val is not None else "Unnamed Product"

                # Clean price
                price_num = 0
                if price_val is not None:
                    try:
                        # Strip currency symbols and commas if it was read as string
                        val_str = str(price_val).replace(',', '').replace('IQD', '').replace('ID', '').replace(' ', '').strip()
                        price_num = int(round(float(val_str)))
                        if price_num > 2147483647:
                            price_num = 2147483647
                    except ValueError:
                        logging.warning(f"Row {row_count+1}: Invalid price value '{price_val}'. Setting to 0.")

                records.append({
                    "barcode": barcode_str,
                    "product_name": name_str,
                    "price_iqd": price_num
                })
                row_count += 1

            logging.info(f"ExcelParser parsed {len(records)} records from {file_path} (auto-mapped).")
            return records
            
        except Exception as e:
            logging.error(f"Error parsing records from {file_path}: {e}")
            raise e
